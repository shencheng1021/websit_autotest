# -*- coding: utf-8 -*-

"""
@author: shencheng
@software: PyCharm
@description: test
@time: 2022/4/13 9:34
"""
import os

import openpyxl


class ExcelUtil:

    def dirname(self):
        return os.path.abspath(os.path.dirname(__file__)).split("common")[0]

    def excel_read(self,filename):
        wb=openpyxl.load_workbook(self.dirname()+r'/data/'+filename+'.xlsx')
        sheet=wb.worksheets[0]
        row=sheet.max_row
        col=sheet.max_column
        row_list=[]
        for r in range(2,row+1):
            col_list=[]
            for col in range(1,col+1):
                datavalue=sheet.cell(row=r,column=col).value
                if datavalue is None:
                    datavalue=''
                    col_list.append(datavalue)
                else:
                    col_list.append(datavalue)
            row_list.append(col_list)

        return row_list

if __name__ == '__main__':
    ExcelUtil().excel_read('login_data')

