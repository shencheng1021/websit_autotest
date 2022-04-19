# -*- coding: utf-8 -*-

"""
@author: shencheng
@software: 砼联数科
@description: 封装一个读写excel的工具包
@time: 2022/4/13 9:34
"""
import os

import openpyxl


class ExcelUtil:

    def get_object_path(self):
        return os.path.abspath(os.path.dirname(__file__)).split("common")[0]

    def read_excel(self,filename,sheetname):
        #加载工作簿
        wb=openpyxl.load_workbook(self.get_object_path()+'data/'+filename+'.xlsx')
        #获得sheet对象
        sheet=wb[sheetname]
        #获取sheet的最大行数和最大列数
        #print(sheet.max_row,sheet.max_column)
        data_list = []
        for row in range(2,sheet.max_row+1):
            row_list = []
            for cols in range(1,sheet.max_column+1):
                if sheet.cell(row,cols).value==None:
                    row_list.append("")
                else:
                    row_list.append(sheet.cell(row,cols).value)
            data_list.append(row_list)
        #print(data_list)
        return data_list

if __name__ == '__main__':
    ExcelUtil().read_excel('login_data')




